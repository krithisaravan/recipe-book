"""
export.py — builds a full Excel export of every table in the app.

Used by:
  - The "Export to Excel" button (manual backup / portability)
  - Anywhere else that wants a point-in-time snapshot of all data

This is a READ-ONLY snapshot — editing the exported file does not write
back to the app's database. It exists so the person always has a way to
get her data out as a portable file, independent of the app staying online.
"""

import io
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from db import (
    get_ingredients, get_packaging, get_recipes,
    get_recipe_ingredients, get_products,
)

HEADER_FILL = PatternFill("solid", fgColor="1A1A1A")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1A1A1A")
NOTE_FONT = Font(name="Calibri", italic=True, size=9, color="6B5C5F")


def _style_sheet(ws, df: pd.DataFrame, title: str, note: str = ""):
    """Apply consistent header styling and column widths to a written sheet."""
    ws.insert_rows(1, amount=3 if note else 2)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if note:
        ws["A2"] = note
        ws["A2"].font = NOTE_FONT

    header_row = 3 if note else 2
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col_name))] +
            [len(str(v)) for v in df[col_name].astype(str).tolist()]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def build_export_workbook() -> bytes:
    """
    Returns the full export as in-memory xlsx bytes, ready for
    st.download_button.
    """
    ingredients = get_ingredients()
    packaging = get_packaging()
    recipes = get_recipes()
    products = get_products()

    ing_df = pd.DataFrame(ingredients) if ingredients else pd.DataFrame(
        columns=["id", "name", "cost", "unit", "cost_per_g", "source"])
    pkg_df = pd.DataFrame(packaging) if packaging else pd.DataFrame(
        columns=["id", "name", "cost_per_pkg", "qty_per_pkg", "cost_per_item", "source"])

    # Recipes summary with base cost rolled up
    recipe_rows = []
    for r in recipes:
        ri = get_recipe_ingredients(r["id"])
        base_cost = sum(x["line_cost"] for x in ri)
        recipe_rows.append({
            "name": r["name"],
            "base_layers": r["base_layers"],
            "base_size_in": r["base_size_in"],
            "base_height_in": r["base_height_in"],
            "base_shape": r["base_shape"],
            "base_cost": round(base_cost, 4),
            "notes": r["notes"],
        })
    recipes_df = pd.DataFrame(recipe_rows) if recipe_rows else pd.DataFrame(
        columns=["name", "base_layers", "base_size_in", "base_height_in",
                 "base_shape", "base_cost", "notes"])

    # Recipe ingredients — flattened, one row per (recipe, ingredient)
    ri_rows = []
    for r in recipes:
        for x in get_recipe_ingredients(r["id"]):
            ri_rows.append({
                "recipe": r["name"],
                "ingredient": x["name"],
                "qty": x["base_qty"],
                "unit": x["unit"],
                "weight_g": x["weight_g"],
                "cost_per_g": x["cost_per_g"],
                "line_cost": x["line_cost"],
            })
    ri_df = pd.DataFrame(ri_rows) if ri_rows else pd.DataFrame(
        columns=["recipe", "ingredient", "qty", "unit", "weight_g", "cost_per_g", "line_cost"])

    products_df = pd.DataFrame(products) if products else pd.DataFrame(
        columns=["name", "recipe_name", "cake_size_in", "cake_shape", "num_layers",
                 "num_cakes", "adjustment_pct", "packaging_name", "packaging_cost"])
    if products:
        products_df = products_df[[
            "name", "recipe_name", "cake_size_in", "cake_shape",
            "num_layers", "num_cakes", "adjustment_pct",
            "packaging_name", "packaging_cost",
        ]].rename(columns={"name": "product", "recipe_name": "recipe"})

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        ing_df.drop(columns=["id"], errors="ignore").to_excel(
            writer, sheet_name="Ingredients", index=False, startrow=2)
        pkg_df.drop(columns=["id"], errors="ignore").to_excel(
            writer, sheet_name="Packaging", index=False, startrow=2)
        recipes_df.to_excel(writer, sheet_name="Recipes", index=False, startrow=2)
        ri_df.to_excel(writer, sheet_name="Recipe Ingredients", index=False, startrow=2)
        products_df.to_excel(writer, sheet_name="Products", index=False, startrow=2)

        wb = writer.book
        timestamp = datetime.now().strftime("%B %d, %Y at %I:%M %p")

        _style_sheet(wb["Ingredients"], ing_df.drop(columns=["id"], errors="ignore"),
                     "Ingredients", f"Exported {timestamp} — read-only snapshot")
        _style_sheet(wb["Packaging"], pkg_df.drop(columns=["id"], errors="ignore"),
                     "Packaging", f"Exported {timestamp} — read-only snapshot")
        _style_sheet(wb["Recipes"], recipes_df,
                     "Recipes", f"Exported {timestamp} — read-only snapshot")
        _style_sheet(wb["Recipe Ingredients"], ri_df,
                     "Recipe Ingredients", "")
        _style_sheet(wb["Products"], products_df,
                     "Products", "")

    buffer.seek(0)
    return buffer.getvalue()
